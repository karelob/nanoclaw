"""
Finance Parsers — KB, Raiffeisen, Revolut bankovní výpisy (PDF), faktury, XLSX.
Read-only — žádná modifikace zdrojových dat.
"""
import re
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional
from decimal import Decimal, InvalidOperation


@dataclass
class Transaction:
    date: str           # DD.MM.YYYY
    description: str    # popis transakce
    counterparty: str   # název protiúčtu
    amount: Decimal     # + příjem, - výdaj
    currency: str = 'CZK'
    vs: str = ''        # variabilní symbol
    note: str = ''      # popis pro mě

    def to_dict(self):
        return {
            'date': self.date,
            'description': self.description,
            'counterparty': self.counterparty,
            'amount': float(self.amount),
            'currency': self.currency,
            'vs': self.vs,
            'note': self.note,
        }


@dataclass
class BankStatement:
    company: str
    account_iban: str
    period_from: str
    period_to: str
    currency: str
    opening_balance: Decimal
    closing_balance: Decimal
    transactions: list[Transaction] = field(default_factory=list)

    @property
    def total_in(self):
        return sum(t.amount for t in self.transactions if t.amount > 0)

    @property
    def total_out(self):
        return sum(t.amount for t in self.transactions if t.amount < 0)

    def summary(self):
        return {
            'company': self.company,
            'iban': self.account_iban,
            'period': f"{self.period_from} – {self.period_to}",
            'currency': self.currency,
            'opening_balance': float(self.opening_balance),
            'closing_balance': float(self.closing_balance),
            'total_in': float(self.total_in),
            'total_out': float(self.total_out),
            'transaction_count': len(self.transactions),
        }


def _parse_amount(s: str) -> Optional[Decimal]:
    """Převede '1 234 567,89' → Decimal('1234567.89')"""
    if not s:
        return None
    s = s.strip().replace('\xa0', '').replace(' ', '').replace(',', '.')
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def parse_kb_bank_statement(pdf_path: Path) -> BankStatement:
    """
    Parsuje Komerční banka periodický výpis z PDF.
    Používá souřadnice slov (pdfplumber) pro spolehlivé oddělení
    částky, VS a popisu — imunní vůči kolizím číslic.

    Sloupce KB výpisu (x-souřadnice):
      Datum       x ≈ 48
      Typ/Popis   90 ≤ x ≤ 410
      VS          410 < x ≤ 510
      Částka      x > 510
    """
    import pdfplumber

    full_text = []
    all_pages_words = []

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                full_text.append(text)
            words = page.extract_words()
            all_pages_words.append(words)

    text = '\n'.join(full_text)

    # --- Metadata z textu ---
    iban_m = re.search(r'IBAN:\s*(CZ\d+)', text)
    period_m = re.search(r'Za období:\s*(\d{2}\.\d{2}\.)\s*[-–]\s*(\d{2}\.\d{2}\.\d{4})', text)
    currency_m = re.search(r'měna:\s*(\w+)', text)
    opening_m = re.search(r'Počáteční zůstatek\s+([\d\s]+,\d{2})', text)
    closing_m = re.search(r'Konečný zůstatek\s+([\d\s]+,\d{2})', text)
    company_m = re.search(r'(BAKER ESTATES|PINEHILL|PINEHOUSE|PINEINVEST|PINEAIR)\s+S\.R\.O\.', text, re.IGNORECASE)

    iban = iban_m.group(1) if iban_m else ''
    period_to = period_m.group(2) if period_m else ''
    period_from = period_m.group(1) + period_to[-4:] if period_m else ''
    currency = currency_m.group(1) if currency_m else 'CZK'
    opening = _parse_amount(opening_m.group(1)) if opening_m else Decimal(0)
    closing = _parse_amount(closing_m.group(1)) if closing_m else Decimal(0)
    company = company_m.group(1).title() if company_m else pdf_path.stem

    # --- Extrahuj "Popis pro mě" poznámky z plného textu ---
    # Formát: "Popis pro mě:\n<text>" nebo "Popis pro mě: <text>" na následujícím řádku
    # Klíčujeme přes (datum, abs(částka)) → note
    notes_by_amount: dict[tuple, str] = {}
    note_blocks = re.finditer(
        r'Popis pro m[eě]:\s*\n(.+?)(?=\n\d{2}\.\d{2}|\nPopis|\nIU|\nOB|\nPokračování|\nKONEČNÝ|\nZpráva pro příjemce|\Z)',
        text, re.DOTALL)
    for nb in note_blocks:
        raw_note = nb.group(1).strip().replace('\n', ' ')
        # Kontext před touto poznámkou (max 500 znaků)
        ctx_start = max(0, nb.start() - 500)
        ctx = text[ctx_start:nb.start()]
        # Datum: poslednímu výskytu DD.MM.YYYY v kontextu
        date_m = re.findall(r'(\d{2}\.\d{2}\.\d{4})', ctx)
        # Částka kdekoliv v kontextu — hledáme všechny, bereme poslední
        amt_m = re.findall(r'(-?\d{1,3}(?:\s\d{3})*,\d{2})', ctx)
        if date_m and amt_m:
            key = (date_m[-1], amt_m[-1].replace(' ', '').replace(',', '.').lstrip('-'))
            notes_by_amount[key] = raw_note

    # --- Parsování transakcí pomocí souřadnic ---
    transactions = []
    date_re = re.compile(r'^\d{2}\.\d{2}\.\d{4}$')
    amount_re = re.compile(r'^-?\d{1,3}(?:\s\d{3})*,\d{2}$')

    for page_words in all_pages_words:
        if not page_words:
            continue

        # Seskup slova do řádků podle y-souřadnice (tolerance 3pt)
        rows = {}
        for w in page_words:
            y = round(w['top'] / 3) * 3
            rows.setdefault(y, []).append(w)

        # Seřaď řádky od shora dolů
        for y_key in sorted(rows.keys()):
            row_words = sorted(rows[y_key], key=lambda w: w['x0'])

            # Řádek transakce začíná datem (první slovo)
            if not row_words:
                continue
            first_word = row_words[0]['text']
            if not date_re.match(first_word):
                continue

            date = first_word

            # Rozděl zbývající slova podle sloupců
            desc_words = [w for w in row_words if 90 <= w['x0'] <= 410]
            vs_words   = [w for w in row_words if 410 < w['x0'] <= 510]
            amt_words  = [w for w in row_words if w['x0'] > 510]

            # Částka — spoj slova a ověř formát
            amount_str = ' '.join(w['text'] for w in amt_words).strip()
            # Odstraň případné nečíselné znaky (záhlaví apod.)
            amount_str = re.sub(r'[^\d\s,\-]', '', amount_str).strip()
            if not amount_re.match(amount_str):
                continue
            amount = _parse_amount(amount_str)
            if amount is None or amount == 0:
                continue

            description_part = ' '.join(w['text'] for w in desc_words).strip()
            vs = ' '.join(w['text'] for w in vs_words).strip()
            # VS musí být čistě číselné
            if vs and not re.match(r'^\d+$', vs):
                vs = ''

            # Filtruj řádky denní zůstatkové tabulky — popis obsahuje datum uvnitř
            if re.search(r'\d{2}\.\d{2}\.\d{4}', description_part):
                continue

            # Dohledej "Popis pro mě" poznámku
            amt_key = str(amount).replace('-', '')  # abs hodnota jako string
            note = notes_by_amount.get((date, amt_key), '')

            transactions.append(Transaction(
                date=date,
                description=description_part,
                counterparty='',
                amount=amount,
                currency=currency,
                vs=vs,
                note=note,
            ))

    return BankStatement(
        company=company,
        account_iban=iban,
        period_from=period_from,
        period_to=period_to,
        currency=currency,
        opening_balance=opening,
        closing_balance=closing,
        transactions=transactions,
    )


def parse_xlsx_cashflow(xlsx_path: Path) -> dict:
    """
    Parsuje cashflow plán z XLSX.
    Vrací dict s klíči: company, rows (list dicts), months.
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    result = {'file': xlsx_path.name, 'sheets': {}}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append(list(row))
        result['sheets'][sheet_name] = rows[:100]  # max 100 řádků

    return result


def parse_pdf_text(pdf_path: Path) -> str:
    """Extrahuje veškerý text z PDF (výsledovky, faktury, apod.)"""
    import pdfplumber
    texts = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                texts.append(t)
    return '\n'.join(texts)


def detect_bank_type(pdf_path: Path) -> str:
    """
    Detekuje typ banky z názvu složky nebo obsahu PDF.
    Vrátí 'kb' | 'rb' | 'revolut' | 'unknown'.
    """
    path_str = str(pdf_path).lower()
    # Složka v cestě
    if '/revolut/' in path_str or '\\revolut\\' in path_str:
        return 'revolut'
    if '/rb/' in path_str or '\\rb\\' in path_str:
        return 'rb'
    # Raiffeisen může být i v cestě s jiným názvem — ověříme text
    try:
        text = parse_pdf_text(pdf_path)[:500]
        if 'raiffeisenbank' in text.lower():
            return 'rb'
        if 'revolut' in text.lower():
            return 'revolut'
        if 'komerční banka' in text.lower() or 'kombczpp' in text.lower():
            return 'kb'
    except Exception:
        pass
    return 'kb'  # default


def parse_statement(pdf_path: Path) -> 'BankStatement':
    """Parsuje výpis automaticky podle banky."""
    bank = detect_bank_type(pdf_path)
    if bank == 'rb':
        return parse_rb_bank_statement(pdf_path)
    if bank == 'revolut':
        return parse_revolut_bank_statement(pdf_path)
    return parse_kb_bank_statement(pdf_path)


def _rb_amount(s: str) -> Optional[Decimal]:
    """Převede Raiffeisen/Revolut číslo '1 234.56' nebo '-1 234.56' na Decimal."""
    s = s.strip().replace('\xa0', '').replace(' ', '')
    # Revolut/RB používají tečku jako desetinný oddělovač
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def parse_rb_bank_statement(pdf_path: Path) -> 'BankStatement':
    """
    Parsuje Raiffeisenbank výpis z PDF.

    Formát transakcí (3 řádky na transakci):
      Řádek 1: D. M. YYYY  (Platba|Poplatek)  Typ  [VS]  AMOUNT CZK
      Řádek 2: D. M. YYYY  [protiucet]  [VS:xxx]  [KS]
      Řádek 3: TxnCode  NázevProtistranY  [Poznámka]
    """
    text = parse_pdf_text(pdf_path)

    # --- Metadata ---
    iban_m = re.search(r'IBAN:\s*(CZ[\d\s]+)', text)
    period_m = re.search(r'za období:\s*(\d{1,2}\. \d{1,2}\. \d{4})\s*[-–]\s*(\d{1,2}\. \d{1,2}\. \d{4})', text)
    currency_m = re.search(r'Výpis z běžného účtu č\.\s*\d+\s+(\w+)', text)
    opening_m = re.search(r'Počáteční zůstatek:\s*([-\d\s.]+)', text)
    closing_m = re.search(r'Konečný zůstatek:\s*([-\d\s.]+)', text)
    name_m = re.search(r'Název účtu:\s*(.+)', text)

    iban = iban_m.group(1).replace(' ', '') if iban_m else ''
    currency = currency_m.group(1).strip() if currency_m else 'CZK'

    def _parse_period_date(s: str) -> str:
        """'D. M. YYYY' → 'DD.MM.YYYY'"""
        parts = re.split(r'\.\s*', s.strip().rstrip('.'))
        if len(parts) >= 3:
            return f"{parts[0].zfill(2)}.{parts[1].zfill(2)}.{parts[2]}"
        return s

    period_from = _parse_period_date(period_m.group(1)) if period_m else ''
    period_to = _parse_period_date(period_m.group(2)) if period_m else ''
    opening = _rb_amount(opening_m.group(1)) if opening_m else Decimal(0)
    closing = _rb_amount(closing_m.group(1)) if closing_m else Decimal(0)
    company = name_m.group(1).strip() if name_m else pdf_path.stem

    # --- Transakce ---
    # Najdeme sekci "Výpis pohybů", ukončenou "Zpráva pro klienta" nebo koncem
    txn_start = text.find('Výpis pohybů')
    txn_end_markers = ['Zpráva pro klienta', 'Vklad na tomto účtu', 'V rámci souhrnné']
    txn_end = len(text)
    for marker in txn_end_markers:
        idx = text.find(marker, txn_start + 1)
        if 0 < idx < txn_end:
            txn_end = idx

    txn_text = text[txn_start:txn_end] if txn_start > 0 else text

    # Řádek 1: začíná datem, patří do kategorie Platba/Poplatek/Konverze
    line1_date_re = re.compile(r'^(\d{1,2})\. (\d{1,2})\. (\d{4})\s+(?:Platba|Poplatek|Konverze)\s')
    # Řádek 3: začíná dlouhým číslem (kód transakce, 7+ číslic)
    line3_re = re.compile(r'^(\d{7,})\s+(.+)$')

    def _extract_rb_amount(line: str, cur: str) -> Optional[Decimal]:
        """
        Extrahuje částku z řádku RB výpisu (odzadu).
        VS (celé číslo bez desetinné tečky) předchází částce — nesmí být zahrnut.
        Pravidla (zprava doleva):
          1. Odeber token s měnou (CZK/USD/EUR)
          2. Najdi poslední token s desetinnou tečkou → základ částky
          3. Absorb nalevo: skupiny PŘESNĚ 3 číslic (tisíce)
          4. Absorb jednoho tokenu: minus+číslice nebo 1–3 číslice (ne VS se 4+ číslicemi)
        """
        tokens = line.split()
        if tokens and tokens[-1] == cur:
            tokens = tokens[:-1]
        # Najdi poslední token s tečkou
        j = next((i for i in range(len(tokens) - 1, -1, -1) if '.' in tokens[i]), -1)
        if j < 0:
            return None
        amt = [tokens[j]]
        k = j - 1
        # Absorb tisícové skupiny (přesně 3 číslice)
        while k >= 0 and re.match(r'^\d{3}$', tokens[k]):
            amt.insert(0, tokens[k])
            k -= 1
        # Absorb leading část: záporný nebo max 3 číslice
        if k >= 0 and re.match(r'^-?\d{1,3}$', tokens[k]):
            amt.insert(0, tokens[k])
        return _rb_amount(''.join(amt))

    transactions = []
    lines = txn_text.split('\n')
    i = 0
    # Přeskočíme hlavičku tabulky
    while i < len(lines) and not line1_date_re.match(lines[i].strip()):
        i += 1

    while i < len(lines):
        line = lines[i].strip()
        if not line:
            i += 1
            continue
        if any(marker in line for marker in txn_end_markers):
            break

        m = line1_date_re.match(line)
        if m:
            day, mon, year = m.group(1), m.group(2), m.group(3)
            date = f"{day.zfill(2)}.{mon.zfill(2)}.{year}"
            cur_m = re.search(r'\b([A-Z]{3})\s*$', line)
            txn_cur = cur_m.group(1) if cur_m else currency
            amount = _extract_rb_amount(line, txn_cur)

            # Řádek 3 = kód transakce + název protistrany
            counterparty = ''
            if i + 2 < len(lines):
                line3 = lines[i + 2].strip()
                m3 = line3_re.match(line3)
                if m3:
                    raw = m3.group(2).strip()
                    raw = re.sub(r'\s+\d+\.?\d*\s*$', '', raw).strip()
                    counterparty = raw

            if amount is not None and amount != 0:
                transactions.append(Transaction(
                    date=date,
                    description=counterparty,
                    counterparty=counterparty,
                    amount=amount,
                    currency=txn_cur,
                ))
            i += 3
        else:
            i += 1

    return BankStatement(
        company=company,
        account_iban=iban,
        period_from=period_from,
        period_to=period_to,
        currency=currency,
        opening_balance=opening or Decimal(0),
        closing_balance=closing or Decimal(0),
        transactions=transactions,
    )


# Měsíce v češtině (Revolut)
_CS_MONTHS = {
    'led': 1, 'úno': 2, 'bře': 3, 'dub': 4, 'kvě': 5, 'čvn': 6,
    'čvc': 7, 'srp': 8, 'zář': 9, 'říj': 10, 'lis': 11, 'pro': 12,
    # Anglické zkratky (starší výpisy)
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
}

# Typ transakce → znaménko (+1 = příjem, -1 = výdaj)
_REVOLUT_SIGN = {
    'CAR': -1,  # platba kartou
    'MOS': -1,  # odeslaná platba
    'EXO': -1,  # směna z
    'FEE': -1,  # poplatek
    'ATM': -1,  # výběr z bankomatu
    'MOR': +1,  # přijatá platba
    'MOA': +1,  # peníze přidány
    'EXI': +1,  # směna za
}
_REVOLUT_TYPES = '|'.join(_REVOLUT_SIGN.keys())


def parse_revolut_bank_statement(pdf_path: Path) -> 'BankStatement':
    """
    Parsuje Revolut Business výpis z PDF.

    Transakční řádek: DD MMM YYYY  TYPE  Popis  [AMOUNT CUR]  BALANCE CUR
    Nebo (USD):       DD MMM YYYY  TYPE  Popis  $AMOUNT  $BALANCE

    Znaménko určeno z typu (CAR/MOS/EXO/FEE/ATM = výdaj, MOR/MOA/EXI = příjem).
    """
    text = parse_pdf_text(pdf_path)

    # --- Metadata ---
    iban_m = re.search(r'IBAN\s+(LT[\w\s]+)', text)
    iban = iban_m.group(1).replace(' ', '') if iban_m else ''

    # Czech ("Měna") or English ("Currency") header; fallback to filename hint (_USD/_EUR/_GBP)
    currency_m = re.search(r'(?:Měna|Currency)\s+([A-Z]{3})', text)
    if currency_m:
        currency = currency_m.group(1)
    else:
        stem = pdf_path.stem.upper()
        currency = 'USD' if '_USD' in stem else 'EUR' if '_EUR' in stem else 'GBP' if '_GBP' in stem else 'CZK'

    # Datum výpisu → určí rok
    gen_m = re.search(r'Vygenerováno\s+(\d+)\.\s+(\w+)\s+(\d{4})', text)
    stmt_year = gen_m.group(3) if gen_m else '2026'

    # Zůstatky (tečka jako desetinný oddělovač)
    opening_m = re.search(
        r'(?:Počáteční zůstatek|Opening balance)\s+(?:[€$£])?(\d[\d\s]*\.?\d*)\s*(?:[A-Z]{3})?', text)
    closing_m = re.search(
        r'(?:Koncový zůstatek|Closing balance)\s+(?:[€$£])?(\d[\d\s]*\.?\d*)\s*(?:[A-Z]{3})?', text)
    opening = _rb_amount(opening_m.group(1)) if opening_m else Decimal(0)
    closing = _rb_amount(closing_m.group(1)) if closing_m else Decimal(0)

    # Název firmy
    name_m = re.search(r'(?:Revolut Bank UAB\s*\n)?\s*(.+?)\s*\nPekařská', text)
    company = name_m.group(1).strip() if name_m else pdf_path.stem

    # --- Transakce ---
    # Sekce transakcí začíná po hlavičce tabulky
    # Hlavička: "Datum (UTC) Popis Odchozí Příchozí Zůstatek"
    # nebo anglicky: "Date (UTC) Description ..."
    hdr_m = re.search(r'Datum \(UTC\)|Date \(UTC\)', text)
    if hdr_m:
        txn_text = text[hdr_m.end():]
    else:
        txn_text = text

    # Konec transakcí: "Typy transakcí" nebo "Transaction types"
    end_m = re.search(r'Typy transakcí|Transaction types', txn_text)
    if end_m:
        txn_text = txn_text[:end_m.start()]

    # Řádek transakce — datum začíná číslem
    # "DD MMM YYYY  TYPE  Description  AMOUNT  BALANCE"
    # nebo "D MMM YYYY ..."
    txn_line_re = re.compile(
        r'^(\d{1,2})\s+([a-záéíóúůčďěňřšťž]{3})\s+(\d{4})\s+'  # datum
        r'(' + _REVOLUT_TYPES + r')\s+'                           # typ
        r'(.+?)\s+'                                               # popis
        r'(?:[€$£])?(\d[\d\s]*\.?\d*)\s+(?:[A-Z]{3}\s+)?'       # částka
        r'(?:[€$£])?(\d[\d\s]*\.?\d*)',                           # zůstatek (zahodíme)
        re.IGNORECASE,
    )
    # Jednodušší regex: jen zachyť datum + typ + konec řádku s čísly
    txn_simple_re = re.compile(
        r'^(\d{1,2})\s+([a-záéíóúůčďěňřšž]{3})\s+(\d{4})\s+(' + _REVOLUT_TYPES + r')\s+(.+)$',
        re.IGNORECASE,
    )
    # Regex pro extrakci částek na konci řádku.
    # Dvě varianty:
    #   1) Prefixovaná měna: $74 982.00  (USD/EUR/GBP — dolar/euro/libra před číslem)
    #   2) Sufixovaná měna: 1 754.70 CZK  (CZK a jiné — kód za číslem, (?<!\d) zabrání
    #      sloučení čísla ID stanice jako "0556" s následnou částkou "179.90")
    amounts_re = re.compile(
        r'(?:[$€£])(\d[\d\s]*\.\d{2})'               # 1) prefixová: $74 982.00
        r'|'
        r'(?<!\d)(\d{1,3}(?:\s\d{3})*\.\d{2})\s+[A-Z]{3}'  # 2) sufixová: 1 754.70 CZK
    )

    transactions = []
    for line in txn_text.split('\n'):
        line = line.strip()
        if not line:
            continue
        m = txn_simple_re.match(line)
        if not m:
            continue
        day_s, mon_s, year_s, txn_type, rest = m.groups()
        mon = _CS_MONTHS.get(mon_s.lower())
        if not mon:
            continue
        date = f"{int(day_s):02d}.{mon:02d}.{year_s}"

        # Extrahuj částky z konce řádku (regex vrací tuple skupin — sloučíme)
        raw_matches = amounts_re.findall(rest)
        nums = [g1 or g2 for g1, g2 in raw_matches]
        if not nums:
            continue

        # Vezmi druhé číslo od konce jako částku (poslední = zůstatek)
        if len(nums) >= 2:
            amt_str = nums[-2]
        else:
            amt_str = nums[-1]

        amount = _rb_amount(amt_str)
        if amount is None or amount == 0:
            continue

        sign = _REVOLUT_SIGN.get(txn_type.upper(), -1)
        amount = amount * sign

        # Popis — odstraň částky z konce (sufixové i prefixové)
        desc = re.sub(r'\s+(?:[$€£])?\d[\d\s]*\.\d{2}\s*(?:[A-Z]{3})?\s*$', '', rest).strip()
        desc = re.sub(r'\s+(?:[$€£])?\d[\d\s]*\.\d{2}\s*(?:[A-Z]{3})?\s*$', '', desc).strip()
        # Odeber typ-prefix pokud přetekl
        desc = desc.strip(' •·–')

        transactions.append(Transaction(
            date=date,
            description=desc,
            counterparty=desc,
            amount=amount,
            currency=currency,
        ))

    # Perioda z prvního/posledního datumu transakcí (Revolut neuvádí period explicitně)
    dates = [t.date for t in transactions]
    p_from = min(dates) if dates else ''
    p_to = max(dates) if dates else ''

    return BankStatement(
        company=company,
        account_iban=iban,
        period_from=p_from,
        period_to=p_to,
        currency=currency,
        opening_balance=opening or Decimal(0),
        closing_balance=closing or Decimal(0),
        transactions=transactions,
    )


if __name__ == '__main__':
    # Test
    stmt = parse_kb_bank_statement(Path('/tmp/baker_bu_sample.pdf'))
    print(stmt.summary())
    print(f"\nPrvní 3 transakce:")
    for t in stmt.transactions[:3]:
        print(f"  {t.date} {t.amount:>12} {t.currency}  {t.description[:40]}  {t.vs}")
